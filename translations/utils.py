import re
from typing import List, Tuple

from django.core.exceptions import ValidationError

import docx
import openpyxl


def extract_sentences_from_text(text: str) -> List[str]:
    """
    Извлекает предложения из текста, используя регулярные выражения
    """
    # Удаляем лишние пробелы и переносы строк
    text = re.sub(r"\s+", " ", text.strip())

    # Паттерн для разделения на предложения
    # Не делим после инициалов вида "С." и учитываем возможные кавычки перед началом следующего предложения
    sentence_pattern = r'(?<!\b[А-ЯЁA-Z]\.)(?<=[.!?])\s+(?=["“”«»]?[А-ЯЁA-Z])'

    # Разбиваем на предложения
    sentences = re.split(sentence_pattern, text)

    # Фильтруем пустые предложения и убираем лишние пробелы
    sentences = [s.strip() for s in sentences if s.strip()]

    # Удаляем дублирующиеся предложения, сохраняя порядок
    seen = set()
    unique_sentences = []
    for sentence in sentences:
        if sentence not in seen:
            seen.add(sentence)
            unique_sentences.append(sentence)

    return unique_sentences


def process_txt_file(file_path: str) -> List[str]:
    """
    Обрабатывает TXT файл и извлекает предложения
    """
    try:
        with open(file_path, "r", encoding="utf-8") as file:
            content = file.read()
        return extract_sentences_from_text(content)
    except UnicodeDecodeError:
        # Пробуем другие кодировки
        encodings = ["cp1251", "latin-1", "iso-8859-1"]
        for encoding in encodings:
            try:
                with open(file_path, "r", encoding=encoding) as file:
                    content = file.read()
                return extract_sentences_from_text(content)
            except UnicodeDecodeError:
                continue
        raise ValidationError("Не удалось прочитать файл. Проверьте кодировку.")


def process_docx_file(file_path: str) -> List[str]:
    """
    Обрабатывает DOCX файл и извлекает абзацы
    """
    try:
        doc = docx.Document(file_path)
        return [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    except Exception as e:
        raise ValidationError(f"Ошибка при чтении DOCX файла: {str(e)}")


def process_xlsx_file(file_path: str) -> List[str]:
    """
    Обрабатывает XLSX файл и извлекает предложения
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        text_content = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            for row in sheet.iter_rows(values_only=True):
                for cell_value in row:
                    if cell_value and isinstance(cell_value, str):
                        text_content.append(str(cell_value))

        # Объединяем весь текст и разбиваем на предложения
        full_text = " ".join(text_content)
        return extract_sentences_from_text(full_text)
    except Exception as e:
        raise ValidationError(f"Ошибка при чтении XLSX файла: {str(e)}")


def process_file(file_path: str, file_extension: str) -> List[str]:
    """
    Основная функция для обработки файлов разных форматов
    """
    file_extension = file_extension.lower()

    if file_extension == ".txt":
        return process_txt_file(file_path)
    elif file_extension == ".docx":
        return process_docx_file(file_path)
    elif file_extension == ".xlsx":
        return process_xlsx_file(file_path)
    else:
        raise ValidationError(f"Неподдерживаемый формат файла: {file_extension}")


def clean_sentence(sentence: str) -> str:
    """
    Очищает предложение от лишних символов и форматирования
    """
    # Убираем лишние пробелы
    sentence = re.sub(r"\s+", " ", sentence.strip())

    # Убираем специальные символы в начале и конце
    # sentence = re.sub(r"^[^\wа-яёА-ЯЁ]*", "", sentence)
    # sentence = re.sub(r"[^\wа-яёА-ЯЁ]*$", "", sentence)

    return sentence


def validate_sentence(sentence: str) -> bool:
    """
    Проверяет, является ли строка валидным предложением
    """
    if not sentence or len(sentence.strip()) < 3:
        return False

    # Проверяем, что предложение содержит буквы
    if not re.search(r"[а-яёА-ЯЁa-zA-Z]", sentence):
        return False

    return True


def extract_and_validate_sentences(file_path: str, file_extension: str) -> List[Tuple[int, str]]:
    """
    Извлекает предложения из файла, валидирует их и возвращает список кортежей (номер, текст)
    """
    sentences = process_file(file_path, file_extension)

    validated_sentences = []
    sentence_number = 1

    for sentence in sentences:
        cleaned_sentence = clean_sentence(sentence)
        if validate_sentence(cleaned_sentence):
            validated_sentences.append((sentence_number, cleaned_sentence))
            sentence_number += 1

    return validated_sentences
